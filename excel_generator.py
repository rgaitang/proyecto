# -*- coding: utf-8 -*-
"""
Generador de Excel con el formato de la plantilla
"BASE LIQUIDACION DE HORAS NUEVA REFORMA 2026".

Estructura generada:
- Hoja REGISTRO   : control de horas por dia (letras de turno) por empleado,
                    con total de horas y VALOR de cada concepto.
- Hoja DATOS      : empleados + matriz letra -> horas por concepto.
- Hoja TURNOS     : tabla de codigos de turno -> descripcion.

El calculo de horas usa la MATRIZ de la hoja DATOS (la configuracion oficial
de la empresa por letra). Los VALORES usan las tarifas de la fila 11 del
REGISTRO (salario 1.750.905 / 210), identicas a la plantilla.
"""
import calendar
import json
import os
import re
from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

_AQUI = os.path.dirname(os.path.abspath(__file__))

# Conceptos en el orden de la plantilla (REGISTRO): (clave_matriz, columna_texto)
CONCEPTOS = [
    ('H.E.D',     'H.E.D'),
    ('H.E.N',     'H.E.N'),
    ('R.NOCTURNO','R.N'),
    ('R.D.D',     'R.D.D'),
    ('R.D.N',     'R.D.N'),
    ('H.E.D.N',   'H.E.D.N'),
    ('H.E.D.D',   'H.E.D.D'),
]

_ARCHIVO_DATOS = os.path.join(_AQUI, 'datos_plantilla.json')


def _cargar_datos():
    with open(_ARCHIVO_DATOS, encoding='utf-8') as f:
        return json.load(f)


def _totales_por_concepto(registros_emp, matriz):
    """A partir de los registros de un empleado (dict {dia: letra}) y la matriz,
    devuelve dict concepto -> horas totales del mes.
    Cada valor puede ser una letra (str) o un objeto RegistroHoras."""
    totales = {c: 0.0 for c, _ in CONCEPTOS}
    for dia, valor in registros_emp.items():
        letra = valor
        if hasattr(valor, 'turno_codigo'):
            letra = valor.turno_codigo
        if not letra:
            continue
        conf = matriz.get(str(letra).strip())
        if not conf:
            continue
        horas = conf.get('horas', {})
        for clave, _ in CONCEPTOS:
            totales[clave] += horas.get(clave, 0) or 0
    return totales


_DIAS_NOMBRE = ['LUNES', 'MARTES', 'MIERCOLES', 'JUEVES', 'VIERNES', 'SABADO', 'DOMINGO']


def _estilo_cabecera(cell):
    cell.font = Font(bold=True, size=10)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.fill = PatternFill('solid', fgColor='D9E1F2')
    cell.border = Border(
        left=Side(style='thin', color='999999'), right=Side(style='thin', color='999999'),
        top=Side(style='thin', color='999999'), bottom=Side(style='thin', color='999999'))


def _crear_hoja_registro(wb, nombre_mes, anio, mes, sucursal_nombre, administrador,
                         empleados, registros, matriz, tarifas):
    dias_mes = calendar.monthrange(anio, mes)[1]
    fechas = [date(anio, mes, d) for d in range(1, dias_mes + 1)]
    n_dias = len(fechas)

    ws = wb.create_sheet('REGISTRO')

    # --- Cabecera ---
    ws.cell(row=1, column=2, value='Registro Y Control De Horas Trabajadas')
    ws.cell(row=2, column=2, value='TURNO')
    ws.cell(row=4, column=2, value='Mes:')
    ws.cell(row=4, column=4, value=nombre_mes.title())
    ws.cell(row=4, column=12, value='Punto De Venta:')
    ws.cell(row=4, column=15, value=sucursal_nombre)
    ws.cell(row=4, column=26, value='Periodo:')
    # Periodo: primer horario de la matriz (o generico)
    periodo = ''
    for conf in matriz.values():
        if conf.get('turno'):
            periodo = conf['turno']
            break
    ws.cell(row=4, column=28, value=periodo)
    ws.cell(row=5, column=2, value='Desde {} Hasta {}'.format(
        fechas[0].strftime('%b-%d'), fechas[-1].strftime('%b-%d')))
    ws.cell(row=6, column=2, value='Año:')
    ws.cell(row=6, column=4, value=anio)
    ws.cell(row=6, column=12, value='Administrador:')
    ws.cell(row=6, column=15, value=administrador)

    # --- Fila de fechas (fila 8) ---
    for i, f in enumerate(fechas):
        ws.cell(row=8, column=5 + i, value=f)

    # --- Encabezado de dias + conceptos (fila 13, como la plantilla) ---
    fila_enc = 13
    ws.cell(row=fila_enc, column=1, value='CEDULA')
    ws.cell(row=fila_enc, column=2, value='')
    ws.cell(row=fila_enc, column=3, value='Asesor')
    for i, f in enumerate(fechas):
        ws.cell(row=fila_enc, column=5 + i, value=_DIAS_NOMBRE[f.weekday()][:3])
    # Columnas de conceptos: (dias empiezan en col 5, terminan en 5+n_dias-1)
    col_conc = 5 + n_dias  # columna del primer concepto (columna del 1er dia + n_dias)
    # En la plantilla la primera columna de concepto es la inmediatamente despues
    # del ultimo dia con una columna vacia de separacion. Asumimos:
    col = 5 + n_dias
    for clave, texto in CONCEPTOS:
        ws.cell(row=fila_enc, column=col, value=texto)
        ws.cell(row=fila_enc, column=col + 1, value='VALOR')
        _estilo_cabecera(ws.cell(row=fila_enc, column=col))
        _estilo_cabecera(ws.cell(row=fila_enc, column=col + 1))
        col += 2
    # VALOR TOTAL
    ws.cell(row=fila_enc, column=col, value='VALOR TOTAL')
    _estilo_cabecera(ws.cell(row=fila_enc, column=col))

    # --- Filas de plantilla de tarifas (filas 11-12) ---
    # Fila 11 = valores unitarios por concepto (tarifas)
    tar_col = 5 + n_dias
    for clave, texto in CONCEPTOS:
        ws.cell(row=11, column=tar_col + 1, value=tarifas.get(clave, 0))
        tar_col += 2

    # --- Filas de empleados ---
    fila = fila_enc + 1   # empieza en fila 14
    for emp in empleados:
        ws.cell(row=fila, column=1, value=emp['cedula'])
        ws.cell(row=fila, column=3, value=emp['nombre'])

        # Letras por dia
        regs = emp['registros']
        for i, f in enumerate(fechas):
            reg = regs.get(f.day)
            letra = reg.turno_codigo if reg else ''
            ws.cell(row=fila, column=5 + i, value=letra if letra else '')

        # Totales por concepto (horas)
        horas_tot = _totales_por_concepto(regs, matriz)

        # Escribir horas y valor
        col = 5 + n_dias
        for clave, texto in CONCEPTOS:
            ws.cell(row=fila, column=col, value=horas_tot.get(clave, 0))
            tarifa = tarifas.get(clave, 0)
            ws.cell(row=fila, column=col + 1, value=round(horas_tot.get(clave, 0) * tarifa, 2))
            col += 2
        total = sum(round(horas_tot.get(clave, 0) * tarifas.get(clave, 0), 2) for clave, _ in CONCEPTOS)
        ws.cell(row=fila, column=col, value=round(total, 2))
        fila += 1

    # formato numerico / ancho
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['C'].width = 30
    return ws


def _crear_hoja_datos(wb, empleados, matriz, sucursal_nombre):
    ws = wb.create_sheet('DATOS')
    # Empleados
    headers = ['CEDULA', 'Nombre', 'Salario', '', '#', 'Punto de Venta']
    for c, h in enumerate(headers, start=1):
        ws.cell(row=3, column=c, value=h)
        _estilo_cabecera(ws.cell(row=3, column=c))
    fila = 4
    for emp in empleados:
        ws.cell(row=fila, column=2, value=emp['cedula'])
        ws.cell(row=fila, column=3, value=emp['nombre'])
        ws.cell(row=fila, column=7, value=sucursal_nombre)
        fila += 1

    # Matriz de conceptos (col 11 en adelante)
    cols_conc = ['CONCEPTO', 'TURNOS', 'H.E.D', 'H.E.N', 'R.NOCTURNO', 'R.D.D', 'R.D.N', 'H.E.D.N', 'H.E.D.D']
    for i, h in enumerate(cols_conc):
        ws.cell(row=3, column=10 + i, value=h)
        _estilo_cabecera(ws.cell(row=3, column=10 + i))
    fila = 4
    for letra in sorted(matriz.keys()):
        conf = matriz[letra]
        ws.cell(row=fila, column=11, value=letra)
        ws.cell(row=fila, column=12, value=conf.get('turno', ''))
        hor = conf.get('horas', {})
        ws.cell(row=fila, column=13, value=hor.get('H.E.D', 0))
        ws.cell(row=fila, column=14, value=hor.get('H.E.N', 0))
        ws.cell(row=fila, column=15, value=hor.get('R.NOCTURNO', 0))
        ws.cell(row=fila, column=16, value=hor.get('R.D.D', 0))
        ws.cell(row=fila, column=17, value=hor.get('R.D.N', 0))
        ws.cell(row=fila, column=18, value=hor.get('H.E.D.N', 0))
        ws.cell(row=fila, column=19, value=hor.get('H.E.D.D', 0))
        fila += 1

    ws.column_dimensions['C'].width = 28
    ws.column_dimensions['L'].width = 32
    return ws


def _crear_hoja_liquidacion(wb, empleados, registros, matriz, tarifas):
    """Crea la hoja 'LIQUIDACION CONCEPTOS': resumen por empleado con las horas
    y el valor de cada concepto + VALOR TOTAL (formato de la plantilla)."""
    ws = wb.create_sheet('LIQUIDACION CONCEPTOS')

    # Cabecera (fila 4) - igual que la plantilla
    headers = ['', 'CEDULA', 'Asesor '] + [x for c in CONCEPTOS for x in (c[1], 'VALOR')] + ['VALOR TOTAL ']
    # Cabecera: B=CEDULA, C=Asesor, D=E H.E.D,V..R=VALOR TOTAL (1-indexed para openpyxl)
    ws.cell(row=4, column=2, value='CEDULA')
    ws.cell(row=4, column=3, value='Asesor ')
    col = 4
    for clave, texto in CONCEPTOS:
        ws.cell(row=4, column=col, value=texto)
        ws.cell(row=4, column=col + 1, value='VALOR')
        _estilo_cabecera(ws.cell(row=4, column=col))
        _estilo_cabecera(ws.cell(row=4, column=col + 1))
        col += 2
    ws.cell(row=4, column=col, value='VALOR TOTAL ')
    _estilo_cabecera(ws.cell(row=4, column=col))

    # Filas de empleados (fila 5 en adelante)
    fila = 5
    for emp in empleados:
        ws.cell(row=fila, column=2, value=emp['cedula'])
        ws.cell(row=fila, column=3, value=emp['nombre'])

        horas_tot = _totales_por_concepto(emp['registros'], matriz)
        col = 4
        total = 0.0
        for clave, texto in CONCEPTOS:
            horas = horas_tot.get(clave, 0)
            valor = round(horas * tarifas.get(clave, 0), 2)
            ws.cell(row=fila, column=col, value=horas)
            ws.cell(row=fila, column=col + 1, value=valor)
            total += valor
            col += 2
        ws.cell(row=fila, column=col, value=round(total, 2))
        fila += 1

    ws.column_dimensions['C'].width = 32
    return ws


def _crear_hoja_turnos(wb, turnos):
    ws = wb.create_sheet('TURNOS')
    ws.cell(row=3, column=2, value='CONCEPTO')
    ws.cell(row=3, column=3, value='TURNOS')
    ws.cell(row=3, column=5, value='CONCEPTO')
    ws.cell(row=3, column=6, value='TURNOS')
    for c in (2, 3, 5, 6):
        _estilo_cabecera(ws.cell(row=3, column=c))
    claves = sorted(turnos.keys())
    mitad = (len(claves) + 1) // 2
    for i in range(mitad):
        fila = 4 + i
        ws.cell(row=fila, column=2, value=claves[i])
        ws.cell(row=fila, column=3, value=turnos[claves[i]])
    for i in range(mitad, len(claves)):
        fila = 4 + (i - mitad)
        ws.cell(row=fila, column=5, value=claves[i])
        ws.cell(row=fila, column=6, value=turnos[claves[i]])
    ws.column_dimensions['C'].width = 36
    ws.column_dimensions['F'].width = 36
    return ws


def generar_archivo_siigo(OUTPUT_DIR, nombre_mes, anio, mes, sucursal_nombre,
                          empleados, registros, novedades, turnos_map):
    """Genera el Excel con el formato de la plantilla BASE LIQUIDACION 2026.

    - `empleados`: empleados del sistema del punto de venta (objetos Empleado).
    - `registros`: dict {empleado_id: {dia_int: RegistroHoras}} del mes.

    Usa los EMPLEADOS DEL SISTEMA como base (los que el admin ve y digita).
    Los empleados que comparten la misma `cedula_real` se agrupan (son la misma
    persona, ej. un nombre corto y su version completa) y sus turnos se
    combinan. El nombre y cedula mostrados son los reales de la hoja DATOS
    cuando estan cargados; si no, los del sistema.
    """
    datos = _cargar_datos()
    matriz = datos['matriz']
    tarifas = datos['tarifas']
    turnos = turnos_map or datos['turnos']

    # Agrupar empleados del sistema por cedula_real (o por su propio nombre si
    # no tienen cedula real). Sumar los registros de los que comparten persona.
    reales = []
    agrupados = {}
    for emp in empleados:
        clave = emp.cedula_real or emp.nombre
        gorup = agrupados.setdefault(clave, {
            'cedula': emp.cedula_real or emp.cedula,
            'nombre': emp.nombre_real or emp.nombre,
            'registros': {},
            'empleado_id': None,
        })
        # combinar turnos (el que tenga letra gana sobre celdas vacias)
        for dia, reg in (registros.get(emp.id) or {}).items():
            if reg is not None and (reg.turno_codigo or reg.turno_codigo == 0):
                gorup['registros'][dia] = reg
        if gorup['empleado_id'] is None:
            gorup['empleado_id'] = emp.id

    reales = list(agrupados.values())
    registros_reales = {emp['nombre']: emp['registros'] for emp in reales}

    administrador = ''
    wb = Workbook()
    wb.remove(wb.active)

    _crear_hoja_registro(wb, nombre_mes, anio, mes, sucursal_nombre, administrador,
                         reales, registros_reales, matriz, tarifas)
    _crear_hoja_datos(wb, reales, matriz, sucursal_nombre)
    _crear_hoja_turnos(wb, turnos)
    _crear_hoja_liquidacion(wb, reales, registros_reales, matriz, tarifas)

    nombre_archivo = f'HORARIOS_{sucursal_nombre.replace(" ", "_")}_{nombre_mes}_{anio}.xlsx'
    ruta = os.path.join(OUTPUT_DIR, nombre_archivo)
    wb.save(ruta)
    return ruta
